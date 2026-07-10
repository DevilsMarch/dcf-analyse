"""Export a DCF as a formula-driven .xlsx workbook (JP-Morgan-style layout).

The workbook is *live*: the Assumptions sheet holds editable input cells and the
DCF sheet is built from Excel formulas that reference them, so the user can keep
tuning the model directly in Excel after download.

Sheets:
    Summary      – headline outputs (both TV methods) + EV→equity bridge
    Assumptions  – all editable inputs (highlighted)
    DCF          – forecast, unlevered FCF, discounting, terminal value
    Sensitivity  – WACC × g and WACC × exit-multiple grids (snapshot values)
"""

from __future__ import annotations

import io

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import ColorScaleRule

from .data import CompanyData
from .assumptions import Assumptions
from .model import DCFResult, run_dcf, _growth_path, _series

# -- styling -----------------------------------------------------------------
NAVY = "1F3864"
LIGHT = "DDEBF7"
INPUT_FILL = "FFF2CC"      # yellow = editable input
CALC_FONT = "1F3864"
GREY = "808080"

_thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)


def _title(ws, cell, text):
    ws[cell] = text
    ws[cell].font = Font(bold=True, size=14, color=NAVY)


def _header(ws, cell, text):
    ws[cell] = text
    ws[cell].font = Font(bold=True, color="FFFFFF")
    ws[cell].fill = PatternFill("solid", fgColor=NAVY)


def _label(ws, cell, text, bold=False, indent=0):
    ws[cell] = text
    ws[cell].font = Font(bold=bold, color="000000")
    if indent:
        ws[cell].alignment = Alignment(indent=indent)


def _input(ws, cell, value, fmt="#,##0.00"):
    ws[cell] = value
    ws[cell].fill = PatternFill("solid", fgColor=INPUT_FILL)
    ws[cell].font = Font(color="7F6000")
    ws[cell].number_format = fmt
    ws[cell].border = BORDER


def _calc(ws, cell, formula, fmt="#,##0.0", bold=False):
    ws[cell] = formula
    ws[cell].font = Font(color=CALC_FONT, bold=bold)
    ws[cell].number_format = fmt


PCT = "0.0%"
PCT2 = "0.00%"
NUM = "#,##0.0"
NUM0 = "#,##0"
MULT = '0.0"x"'


def build_workbook(data: CompanyData, a: Assumptions,
                   result: DCFResult | None = None,
                   mid_year: bool = True, model_values: dict | None = None) -> Workbook:
    if result is None:
        result = run_dcf(data, a, mid_year=mid_year)

    wb = Workbook()

    # ---- Assumptions sheet (built first; other sheets reference it) --------
    A = wb.active
    A.title = "Assumptions"
    _build_assumptions(A, data, a)

    # ---- DCF sheet --------------------------------------------------------
    D = wb.create_sheet("DCF")
    _build_dcf(D, data, a, result, mid_year)

    # ---- Charts sheet -----------------------------------------------------
    C = wb.create_sheet("Charts")
    _build_charts(C, D, data, a)

    # ---- Summary sheet ----------------------------------------------------
    S = wb.create_sheet("Summary")
    _build_summary(S, data, a)
    wb.move_sheet("Summary", -3)  # put Summary first

    # ---- Sensitivity ------------------------------------------------------
    Z = wb.create_sheet("Sensitivity")
    _build_sensitivity(Z, data, a, result)

    # ---- Models overview (optional) ---------------------------------------
    if model_values:
        M = wb.create_sheet("Modelle")
        _build_models(M, data, model_values)

    wb.active = 0
    return wb


# Cell map for the Assumptions sheet (label row -> value cell in column C).
# Kept as a module-level dict so the DCF sheet can reference by name.
AR = {
    "company": 3, "currency": 4, "price": 5, "price_to_major": 6, "shares": 7,
    "net_debt": 8, "minorities": 9, "pension": 10, "associates": 11,
    "forecast_years": 14, "init_growth": 15, "term_growth": 16, "ebitda_margin": 17,
    "da_pct": 18, "capex_pct": 19, "nwc_pct": 20, "tax_rate": 21,
    "risk_free": 24, "erp": 25, "beta": 26, "cost_equity": 27, "cost_debt": 28,
    "equity_w": 29, "debt_w": 30, "wacc": 31,
    "perp_growth": 34, "exit_mult": 35,
}


def _ac(key: str) -> str:
    """Absolute reference to an Assumptions input cell, e.g. Assumptions!$C$15."""
    return f"Assumptions!$C${AR[key]}"


def _build_assumptions(ws, data: CompanyData, a: Assumptions):
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 34
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 40

    _title(ws, "B1", f"DCF Assumptions — {data.name} ({data.ticker})")
    ws["B1"].font = Font(bold=True, size=14, color=NAVY)

    _header(ws, "B2", "Company & market")
    _header(ws, "C2", "Value")
    _label(ws, f"B{AR['company']}", "Company / ticker")
    ws[f"C{AR['company']}"] = f"{data.name} ({data.ticker})"
    _label(ws, f"B{AR['currency']}", "Reporting currency")
    ws[f"C{AR['currency']}"] = data.reporting_currency
    _label(ws, f"B{AR['price']}", f"Current share price ({data.price_currency})")
    _input(ws, f"C{AR['price']}", round(data.price, 4))
    _label(ws, f"B{AR['price_to_major']}", "Price units per reporting unit")
    _input(ws, f"C{AR['price_to_major']}", data.price_to_major, fmt="0")
    ws[f"D{AR['price_to_major']}"] = "100 if price quoted in pence/cents, else 1"
    ws[f"D{AR['price_to_major']}"].font = Font(italic=True, color=GREY, size=9)
    _label(ws, f"B{AR['shares']}", "Shares outstanding (m)")
    _input(ws, f"C{AR['shares']}", round(data.shares_out, 3), fmt=NUM)
    _label(ws, f"B{AR['net_debt']}", "Net debt / (cash) (m)")
    _nd = a.net_debt_override if a.net_debt_override is not None else data.net_debt
    _input(ws, f"C{AR['net_debt']}", round(_nd, 1), fmt=NUM)
    _label(ws, f"B{AR['minorities']}", "Minority interests (m)")
    _input(ws, f"C{AR['minorities']}", a.minority_interests, fmt=NUM)
    _label(ws, f"B{AR['pension']}", "Pension liability (m)")
    _input(ws, f"C{AR['pension']}", a.pension_liability, fmt=NUM)
    _label(ws, f"B{AR['associates']}", "Investment in associates (m)")
    _input(ws, f"C{AR['associates']}", a.associates, fmt=NUM)

    _header(ws, "B13", "Operating forecast")
    _header(ws, "C13", "Value")
    _label(ws, f"B{AR['forecast_years']}", "Forecast years")
    _input(ws, f"C{AR['forecast_years']}", a.forecast_years, fmt="0")
    _label(ws, f"B{AR['init_growth']}", "Revenue growth (per year)")
    ws[f"C{AR['init_growth']}"] = "→ DCF!row 5"
    ws[f"C{AR['init_growth']}"].font = Font(italic=True, color=GREY, size=9)
    _label(ws, f"B{AR['term_growth']}", f"Growth source")
    ws[f"C{AR['term_growth']}"] = a.growth_source
    ws[f"C{AR['term_growth']}"].font = Font(italic=True, color=GREY, size=9)
    # Operating margins & ratios are now per-year (editable on the DCF sheet).
    def _seed_note(r, label, value):
        _label(ws, f"B{r}", label)
        ws[f"C{r}"] = f"{value:.1%} → DCF"
        ws[f"C{r}"].font = Font(italic=True, color=GREY, size=9)

    _seed_note(AR['ebitda_margin'], "EBITDA margin (per year)", a.ebitda_margin)
    _seed_note(AR['da_pct'], "D&A % (per year)", a.da_pct_revenue)
    _seed_note(AR['capex_pct'], "Capex % (per year)", a.capex_pct_revenue)
    _seed_note(AR['nwc_pct'], "Δ NWC % (per year)", a.nwc_pct_revenue_change)
    _label(ws, f"B{AR['tax_rate']}", "Tax rate")
    _input(ws, f"C{AR['tax_rate']}", a.tax_rate, fmt=PCT)

    _header(ws, "B23", "WACC build-up")
    _header(ws, "C23", "Value")
    _label(ws, f"B{AR['risk_free']}", "Risk-free rate")
    _input(ws, f"C{AR['risk_free']}", a.risk_free, fmt=PCT2)
    _label(ws, f"B{AR['erp']}", "Equity risk premium")
    _input(ws, f"C{AR['erp']}", a.equity_risk_premium, fmt=PCT2)
    _label(ws, f"B{AR['beta']}", "Beta")
    _input(ws, f"C{AR['beta']}", a.beta, fmt="0.00")
    _label(ws, f"B{AR['cost_equity']}", "Cost of equity", bold=True)
    _calc(ws, f"C{AR['cost_equity']}",
          f"={_ac('risk_free')}+{_ac('beta')}*{_ac('erp')}", fmt=PCT2, bold=True)
    _label(ws, f"B{AR['cost_debt']}", "Pre-tax cost of debt")
    _input(ws, f"C{AR['cost_debt']}", a.pretax_cost_of_debt, fmt=PCT2)
    _label(ws, f"B{AR['equity_w']}", "Equity weight E/(D+E)")
    _input(ws, f"C{AR['equity_w']}", a.equity_weight, fmt=PCT)
    _label(ws, f"B{AR['debt_w']}", "Debt weight D/(D+E)")
    _input(ws, f"C{AR['debt_w']}", a.debt_weight, fmt=PCT)
    _label(ws, f"B{AR['wacc']}", "WACC", bold=True)
    _calc(ws, f"C{AR['wacc']}",
          f"={_ac('equity_w')}*{_ac('cost_equity')}+{_ac('debt_w')}*{_ac('cost_debt')}*(1-{_ac('tax_rate')})",
          fmt=PCT2, bold=True)
    ws[f"C{AR['wacc']}"].fill = PatternFill("solid", fgColor=LIGHT)

    _header(ws, "B33", "Terminal value")
    _header(ws, "C33", "Value")
    _label(ws, f"B{AR['perp_growth']}", "Perpetuity growth rate")
    _input(ws, f"C{AR['perp_growth']}", a.perpetuity_growth, fmt=PCT2)
    _label(ws, f"B{AR['exit_mult']}", "Exit EBITDA multiple")
    _input(ws, f"C{AR['exit_mult']}", a.exit_ebitda_multiple, fmt=MULT)

    note = ws["B37"]
    note.value = "Yellow cells are editable inputs. Blue cells are calculated — change inputs and Excel recomputes the DCF."
    note.font = Font(italic=True, color=GREY, size=9)


def _build_dcf(ws, data: CompanyData, a: Assumptions, r: DCFResult, mid_year: bool):
    ws.column_dimensions["B"].width = 32
    ws.column_dimensions["C"].width = 12
    n = a.forecast_years
    base_col = 4  # column D = last historical (year 0)
    last_col_idx = base_col + n
    last_col = get_column_letter(last_col_idx)

    def col(i):  # forecast year i (1..n) -> column letter
        return get_column_letter(base_col + i)

    _title(ws, "B1", f"Discounted Cash Flow — {data.name} ({data.ticker})")

    # header rows: row 3 = forecast year index, row 4 = fiscal year label
    _label(ws, "B3", "Forecast year", bold=True)
    _label(ws, "B4", "Fiscal year", bold=True)
    ws["D3"] = 0
    ws["D3"].font = Font(bold=True, italic=True, color=GREY)
    base_year = data.years[-1] if data.years else 0
    ws["D4"] = base_year
    ws["D4"].font = Font(bold=True, italic=True, color=GREY)
    for i in range(1, n + 1):
        c = f"{col(i)}3"
        ws[c] = i
        ws[c].font = Font(bold=True, color=NAVY)
        ws[c].alignment = Alignment(horizontal="center")
        cy = f"{col(i)}4"
        ws[cy] = base_year + i
        ws[cy].font = Font(bold=True, color=NAVY)
        ws[cy].alignment = Alignment(horizontal="center")
    ws[f"{last_col}3"].alignment = Alignment(horizontal="center")
    tv_col = get_column_letter(last_col_idx + 1)
    ws[f"{tv_col}3"] = "TV"
    ws[f"{tv_col}3"].font = Font(bold=True, color=NAVY)
    ws.freeze_panes = "C5"

    rows = {
        "growth": 5, "revenue": 6, "margin": 7, "ebitda": 8, "dapct": 9, "da": 10,
        "ebit": 11, "taxrate": 12, "tax": 13, "ebiat": 14, "capexpct": 15, "capex": 16,
        "nwcpct": 17, "nwc": 18, "ufcf": 20, "period": 22, "disc": 23, "pv": 24,
    }

    _label(ws, f"B{rows['growth']}", "Revenue growth %")
    _label(ws, f"B{rows['revenue']}", "Revenue", bold=True)
    _label(ws, f"B{rows['margin']}", "EBITDA margin %")
    _label(ws, f"B{rows['ebitda']}", "EBITDA")
    _label(ws, f"B{rows['dapct']}", "D&A % of revenue")
    _label(ws, f"B{rows['da']}", "D&A")
    _label(ws, f"B{rows['ebit']}", "EBIT")
    _label(ws, f"B{rows['taxrate']}", "Tax rate %")
    _label(ws, f"B{rows['tax']}", "Tax on EBIT")
    _label(ws, f"B{rows['ebiat']}", "EBIAT")
    _label(ws, f"B{rows['capexpct']}", "Capex % of revenue")
    _label(ws, f"B{rows['capex']}", "Capex")
    _label(ws, f"B{rows['nwcpct']}", "Δ NWC % of revenue")
    _label(ws, f"B{rows['nwc']}", "Δ Working capital")
    _label(ws, f"B{rows['ufcf']}", "Unlevered FCF", bold=True)
    _label(ws, f"B{rows['period']}", "Discount period (yrs)")
    _label(ws, f"B{rows['disc']}", "Discount factor")
    _label(ws, f"B{rows['pv']}", "PV of FCF", bold=True)

    # base year (col D): last historical revenue only, as anchor
    _calc(ws, f"D{rows['revenue']}", data.last_revenue, fmt=NUM0, bold=True)

    # explicit per-year driver series (editable yellow input cells)
    gpath = _growth_path(a)
    mpath = _series(a.ebitda_margin_path, a.ebitda_margin, n)
    dpath = _series(a.da_pct_path, a.da_pct_revenue, n)
    cpath = _series(a.capex_pct_path, a.capex_pct_revenue, n)
    wpath = _series(a.nwc_pct_path, a.nwc_pct_revenue_change, n)
    half = 0.5 if mid_year else 0.0
    for i in range(1, n + 1):
        c = col(i)
        prev = get_column_letter(base_col + i - 1)
        yr = f"{c}3"
        _input(ws, f"{c}{rows['growth']}", round(gpath[i - 1], 4), fmt=PCT)
        _calc(ws, f"{c}{rows['revenue']}", f"={prev}{rows['revenue']}*(1+{c}{rows['growth']})", fmt=NUM0, bold=True)
        _input(ws, f"{c}{rows['margin']}", round(mpath[i - 1], 4), fmt=PCT)
        _calc(ws, f"{c}{rows['ebitda']}", f"={c}{rows['revenue']}*{c}{rows['margin']}", fmt=NUM)
        _input(ws, f"{c}{rows['dapct']}", round(dpath[i - 1], 4), fmt=PCT)
        _calc(ws, f"{c}{rows['da']}", f"={c}{rows['revenue']}*{c}{rows['dapct']}", fmt=NUM)
        _calc(ws, f"{c}{rows['ebit']}", f"={c}{rows['ebitda']}-{c}{rows['da']}", fmt=NUM)
        ws[f"{c}{rows['taxrate']}"] = f"={_ac('tax_rate')}"
        ws[f"{c}{rows['taxrate']}"].number_format = PCT
        ws[f"{c}{rows['taxrate']}"].font = Font(color=CALC_FONT)
        _calc(ws, f"{c}{rows['tax']}", f"=-MAX({c}{rows['ebit']},0)*{c}{rows['taxrate']}", fmt=NUM)
        _calc(ws, f"{c}{rows['ebiat']}", f"={c}{rows['ebit']}+{c}{rows['tax']}", fmt=NUM)
        _input(ws, f"{c}{rows['capexpct']}", round(cpath[i - 1], 4), fmt=PCT)
        _calc(ws, f"{c}{rows['capex']}", f"=-{c}{rows['revenue']}*{c}{rows['capexpct']}", fmt=NUM)
        _input(ws, f"{c}{rows['nwcpct']}", round(wpath[i - 1], 4), fmt=PCT)
        _calc(ws, f"{c}{rows['nwc']}", f"={c}{rows['revenue']}*{c}{rows['nwcpct']}", fmt=NUM)
        _calc(ws, f"{c}{rows['ufcf']}", f"={c}{rows['ebiat']}+{c}{rows['da']}+{c}{rows['capex']}+{c}{rows['nwc']}", fmt=NUM, bold=True)

        _calc(ws, f"{c}{rows['period']}", f"={yr}-{half}", fmt="0.0")
        _calc(ws, f"{c}{rows['disc']}", f"=1/(1+{_ac('wacc')})^{c}{rows['period']}", fmt="0.000")
        _calc(ws, f"{c}{rows['pv']}", f"={c}{rows['ufcf']}*{c}{rows['disc']}", fmt=NUM, bold=True)

    # ---- valuation block --------------------------------------------------
    first_fc = col(1)
    br = 27
    _header(ws, f"B{br}", "Valuation")
    _header(ws, f"C{br}", "Perpetuity")
    _header(ws, f"D{br}", "Exit multiple")

    _label(ws, f"B{br+1}", "PV of forecast FCF")
    _calc(ws, f"C{br+1}", f"=SUM({first_fc}{rows['pv']}:{last_col}{rows['pv']})", fmt=NUM0)
    _calc(ws, f"D{br+1}", f"=C{br+1}", fmt=NUM0)

    _label(ws, f"B{br+2}", "Terminal value (undiscounted)")
    _calc(ws, f"C{br+2}",
          f"={last_col}{rows['ufcf']}*(1+{_ac('perp_growth')})/({_ac('wacc')}-{_ac('perp_growth')})", fmt=NUM0)
    _calc(ws, f"D{br+2}", f"={last_col}{rows['ebitda']}*{_ac('exit_mult')}", fmt=NUM0)

    _label(ws, f"B{br+3}", "PV of terminal value")
    _calc(ws, f"C{br+3}", f"=C{br+2}*{last_col}{rows['disc']}", fmt=NUM0)
    _calc(ws, f"D{br+3}", f"=D{br+2}*{last_col}{rows['disc']}", fmt=NUM0)

    _label(ws, f"B{br+4}", "Enterprise value", bold=True)
    _calc(ws, f"C{br+4}", f"=C{br+1}+C{br+3}", fmt=NUM0, bold=True)
    _calc(ws, f"D{br+4}", f"=D{br+1}+D{br+3}", fmt=NUM0, bold=True)

    _label(ws, f"B{br+5}", "less: Net debt")
    _calc(ws, f"C{br+5}", f"=-{_ac('net_debt')}", fmt=NUM0)
    _calc(ws, f"D{br+5}", f"=-{_ac('net_debt')}", fmt=NUM0)
    _label(ws, f"B{br+6}", "less: Minorities/pension/associates")
    _calc(ws, f"C{br+6}", f"=-({_ac('minorities')}+{_ac('pension')}+{_ac('associates')})", fmt=NUM0)
    _calc(ws, f"D{br+6}", f"=-({_ac('minorities')}+{_ac('pension')}+{_ac('associates')})", fmt=NUM0)

    _label(ws, f"B{br+7}", "Equity value", bold=True)
    _calc(ws, f"C{br+7}", f"=C{br+4}+C{br+5}+C{br+6}", fmt=NUM0, bold=True)
    _calc(ws, f"D{br+7}", f"=D{br+4}+D{br+5}+D{br+6}", fmt=NUM0, bold=True)

    _label(ws, f"B{br+8}", "Implied share price", bold=True)
    _calc(ws, f"C{br+8}", f"=C{br+7}/{_ac('shares')}*{_ac('price_to_major')}", fmt=NUM, bold=True)
    _calc(ws, f"D{br+8}", f"=D{br+7}/{_ac('shares')}*{_ac('price_to_major')}", fmt=NUM, bold=True)
    ws[f"C{br+8}"].fill = PatternFill("solid", fgColor=LIGHT)
    ws[f"D{br+8}"].fill = PatternFill("solid", fgColor=LIGHT)

    _label(ws, f"B{br+9}", "Premium / (discount) to current")
    _calc(ws, f"C{br+9}", f"=C{br+8}/{_ac('price')}-1", fmt=PCT)
    _calc(ws, f"D{br+9}", f"=D{br+8}/{_ac('price')}-1", fmt=PCT)

    # store the block anchor for Summary references
    ws._dcf_block_row = br  # type: ignore[attr-defined]


def _build_summary(ws, data: CompanyData, a: Assumptions):
    ws.column_dimensions["B"].width = 36
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 16
    br = 27  # matches _build_dcf block row

    _title(ws, "B1", f"DCF Summary — {data.name} ({data.ticker})")
    ws["B2"] = "Semi-automatic DCF · data via Yahoo Finance · assumptions editable on the Assumptions sheet"
    ws["B2"].font = Font(italic=True, color=GREY, size=9)

    _header(ws, "B4", "Output")
    _header(ws, "C4", "Perpetuity growth")
    _header(ws, "D4", "Exit multiple")

    labels = [
        ("Enterprise value (m)", br + 4, NUM0),
        ("Equity value (m)", br + 7, NUM0),
        ("Implied share price", br + 8, NUM),
        ("Premium / (discount) to current", br + 9, PCT),
    ]
    row = 5
    for text, dcf_row, fmt in labels:
        _label(ws, f"B{row}", text, bold=(text == "Implied share price"))
        _calc(ws, f"C{row}", f"=DCF!C{dcf_row}", fmt=fmt, bold=(text == "Implied share price"))
        _calc(ws, f"D{row}", f"=DCF!D{dcf_row}", fmt=fmt, bold=(text == "Implied share price"))
        row += 1

    _label(ws, f"B{row+1}", f"Current share price ({data.price_currency})", bold=True)
    _calc(ws, f"C{row+1}", f"={_ac('price')}", fmt=NUM, bold=True)
    _label(ws, f"B{row+2}", "WACC")
    _calc(ws, f"C{row+2}", f"={_ac('wacc')}", fmt=PCT2)
    _label(ws, f"B{row+3}", "Perpetuity growth / exit multiple")
    _calc(ws, f"C{row+3}", f"={_ac('perp_growth')}", fmt=PCT2)
    _calc(ws, f"D{row+3}", f"={_ac('exit_mult')}", fmt=MULT)


def _build_sensitivity(ws, data: CompanyData, a: Assumptions, r: DCFResult):
    ws.column_dimensions["B"].width = 14
    _title(ws, "B1", "Sensitivity analysis (snapshot)")
    ws["B2"] = "Implied share price. Recompute in the app after changing assumptions for live values."
    ws["B2"].font = Font(italic=True, color=GREY, size=9)

    s = r.sensitivity

    def grid(top, title, col_vals, col_fmt, price_grid, col_header):
        _header(ws, f"B{top}", title)
        ws[f"C{top}"] = col_header
        ws[f"C{top}"].font = Font(bold=True, italic=True, color=NAVY)
        # column headers (g or multiple)
        for j, cv in enumerate(col_vals):
            cell = ws.cell(row=top + 1, column=3 + j)
            cell.value = cv
            cell.number_format = col_fmt
            cell.font = Font(bold=True, color=NAVY)
            cell.alignment = Alignment(horizontal="center")
        ws.cell(row=top + 1, column=2, value="WACC ↓").font = Font(bold=True, italic=True, color=NAVY)
        for i, w in enumerate(s["waccs"]):
            rr = top + 2 + i
            wc = ws.cell(row=rr, column=2, value=w)
            wc.number_format = PCT2
            wc.font = Font(bold=True, color=NAVY)
            for j, _ in enumerate(col_vals):
                cc = ws.cell(row=rr, column=3 + j, value=round(price_grid[i][j], 1))
                cc.number_format = NUM
                cc.border = BORDER
        # colour-scale conditional formatting across the data cells
        first = f"{get_column_letter(3)}{top + 2}"
        last = f"{get_column_letter(2 + len(col_vals))}{top + 1 + len(s['waccs'])}"
        ws.conditional_formatting.add(
            f"{first}:{last}",
            ColorScaleRule(start_type="min", start_color="C0504D",
                           mid_type="percentile", mid_value=50, mid_color="FFEB84",
                           end_type="max", end_color="4C9F70"),
        )

    grid(4, "Price vs WACC × perpetuity growth", s["growths"], PCT2, s["price_growth"], "Perpetuity growth →")
    grid(4 + len(s["waccs"]) + 4, "Price vs WACC × exit multiple", s["multiples"], MULT, s["price_multiple"], "Exit multiple →")


def _build_charts(ws, dcf_ws, data: CompanyData, a: Assumptions):
    """A small formula-linked data table on the left, native Excel charts on the
    right. Charts reference the local table so they render everywhere."""
    n = a.forecast_years
    base_col = 4

    def dcol(i):
        return get_column_letter(base_col + i)

    _title(ws, "A1", f"Charts — {data.name} ({data.ticker})")
    ws["A2"] = "Werte sind mit dem DCF-Blatt verknüpft und aktualisieren sich automatisch."
    ws["A2"].font = Font(italic=True, color=GREY, size=9)

    headers = ["Jahr", "Umsatz", "EBITDA", "Unlev. FCF", "PV FCF", "Wachstum"]
    hdr = 4
    for j, h in enumerate(headers):
        cell = ws.cell(row=hdr, column=1 + j, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=NAVY)
    for i in range(1, n + 1):
        r = hdr + i
        c = dcol(i)
        ws.cell(r, 1, value=f"=DCF!{c}4")
        ws.cell(r, 2, value=f"=DCF!{c}6").number_format = NUM0     # revenue
        ws.cell(r, 3, value=f"=DCF!{c}8").number_format = NUM0     # EBITDA
        ws.cell(r, 4, value=f"=DCF!{c}20").number_format = NUM0    # unlevered FCF
        ws.cell(r, 5, value=f"=DCF!{c}24").number_format = NUM0    # PV of FCF
        ws.cell(r, 6, value=f"=DCF!{c}5").number_format = PCT      # growth
    end = hdr + n
    ws.column_dimensions["A"].width = 10

    cats = Reference(ws, min_col=1, min_row=hdr + 1, max_row=end)
    ccy = data.reporting_currency

    bar = BarChart()
    bar.type = "col"; bar.grouping = "clustered"; bar.title = "Umsatz & EBITDA"
    bar.height = 8; bar.width = 17; bar.style = 10
    bar.add_data(Reference(ws, min_col=2, max_col=3, min_row=hdr, max_row=end), titles_from_data=True)
    bar.set_categories(cats)
    bar.y_axis.title = f"Mio {ccy}"; bar.x_axis.title = "Jahr"
    ws.add_chart(bar, "H4")

    bar2 = BarChart()
    bar2.type = "col"; bar2.grouping = "clustered"; bar2.title = "Free Cash Flow & Barwert"
    bar2.height = 8; bar2.width = 17; bar2.style = 12
    bar2.add_data(Reference(ws, min_col=4, max_col=5, min_row=hdr, max_row=end), titles_from_data=True)
    bar2.set_categories(cats)
    bar2.y_axis.title = f"Mio {ccy}"; bar2.x_axis.title = "Jahr"
    ws.add_chart(bar2, "H21")

    ln = LineChart()
    ln.title = "Umsatzwachstum (%)"; ln.height = 8; ln.width = 17; ln.style = 13
    ln.add_data(Reference(ws, min_col=6, max_col=6, min_row=hdr, max_row=end), titles_from_data=True)
    ln.set_categories(cats)
    ln.y_axis.numFmt = "0%"; ln.x_axis.title = "Jahr"
    ws.add_chart(ln, "H38")

    # Pie: PV of FCF vs PV of Terminal Value
    ws.cell(end + 3, 1, value="PV der FCF")
    ws.cell(end + 3, 2, value="=DCF!C28").number_format = NUM0
    ws.cell(end + 4, 1, value="PV Terminal Value")
    ws.cell(end + 4, 2, value="=DCF!C30").number_format = NUM0
    pie = PieChart()
    pie.title = "EV-Zusammensetzung"; pie.height = 8; pie.width = 11
    pie.add_data(Reference(ws, min_col=2, min_row=end + 3, max_row=end + 4), titles_from_data=False)
    pie.set_categories(Reference(ws, min_col=1, min_row=end + 3, max_row=end + 4))
    pie.dataLabels = DataLabelList(); pie.dataLabels.showPercent = True
    ws.add_chart(pie, "H55")


def _build_models(ws, data: CompanyData, model_values: dict):
    """Overview of all valuation methods and their implied prices."""
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 16
    _title(ws, "B1", f"Bewertungsmethoden — {data.name} ({data.ticker})")
    ws["B2"] = f"Impliziter Kurs je Methode · Marktkurs {data.price:,.2f} {data.price_currency}"
    ws["B2"].font = Font(italic=True, color=GREY, size=9)

    _header(ws, "B4", "Methode")
    _header(ws, "C4", f"Impliziter Kurs ({data.price_currency})")
    _header(ws, "D4", "Prämie/(Disc.)")
    row = 5
    for name, price in model_values.items():
        if price is None:
            continue
        _label(ws, f"B{row}", name, bold=name.startswith("DCF"))
        _calc(ws, f"C{row}", round(float(price), 2), fmt=NUM, bold=name.startswith("DCF"))
        prem = price / data.price - 1 if data.price else 0
        _calc(ws, f"D{row}", round(prem, 4), fmt=PCT)
        row += 1

    _label(ws, f"B{row + 1}", "Aktueller Marktkurs", bold=True)
    _calc(ws, f"C{row + 1}", round(data.price, 2), fmt=NUM, bold=True)
    ws[f"C{row + 1}"].fill = PatternFill("solid", fgColor=LIGHT)


def workbook_bytes(data: CompanyData, a: Assumptions,
                   result: DCFResult | None = None,
                   model_values: dict | None = None) -> bytes:
    """Return the .xlsx as bytes (for Streamlit download_button)."""
    wb = build_workbook(data, a, result, model_values=model_values)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
